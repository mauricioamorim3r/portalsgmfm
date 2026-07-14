"""Gera uma migração D1 a partir de uma campanha real de calibração MPFM.

Lê o Excel de campanha (formato `MPFM_P4_V2_3_Graficos(2).xlsx`, 27 abas)
embutido no pacote `Portal_MPFM_Riser_P4_v1.zip` e converte as abas de
entrada manual (`IN_*`) e o contrato da campanha (`01_CAMPANHA`) em SQL
INSERT para as tabelas calibration_* — sem inferir nem completar nenhum
campo que a fonte deixou em branco.
"""

from __future__ import annotations

import argparse
import hashlib
import io
import zipfile
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


def sql_value(value):
    if value is None or value == "":
        return "NULL"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, (int, float)):
        return str(value) if value == value else "NULL"
    if isinstance(value, (datetime, date)):
        return "'" + value.strftime("%Y-%m-%dT%H:%M:%S") + "'"
    return "'" + str(value).replace("'", "''") + "'"


def number(value):
    if value in (None, ""):
        return None
    try:
        result = float(value)
        return result if result == result else None
    except (TypeError, ValueError):
        return None


def load_campaign_workbook(zip_path: Path):
    with zipfile.ZipFile(zip_path) as archive:
        [xlsx_name] = [name for name in archive.namelist() if name.startswith("upload/") and name.endswith(".xlsx")]
        xlsx_bytes = archive.read(xlsx_name)
    sha256 = hashlib.sha256(xlsx_bytes).hexdigest()
    return load_workbook(io.BytesIO(xlsx_bytes), data_only=True), Path(xlsx_name).name, sha256


def cell(ws, row, col=2):
    return ws.cell(row=row, column=col).value


def add_campaign(wb, file_name: str, sha256: str, statements: list[str]) -> str:
    ws = wb["01_CAMPANHA"]
    campaign_id = cell(ws, 5)
    values = {
        "campaign_id": campaign_id,
        "revision": cell(ws, 6),
        "nature": cell(ws, 7),
        "asset": cell(ws, 8),
        "well": cell(ws, 9),
        "tag": cell(ws, 10),
        "serial": cell(ws, 11),
        "reference_tag": cell(ws, 13),
        "start_at": cell(ws, 14),
        "end_at": cell(ws, 15),
        "post_start_at": cell(ws, 17),
        "post_end_at": cell(ws, 18),
        "pb_barg": number(cell(ws, 20)),
        "hc_limit_pct": number(cell(ws, 24)),
        "total_limit_pct": number(cell(ws, 25)),
        "pvt_limit_pct": number(cell(ws, 26)),
        "k_min": number(cell(ws, 27)),
        "k_max": number(cell(ws, 28)),
        "min_records": number(cell(ws, 29)),
        "pvt_months": number(cell(ws, 30)),
        "responsible": cell(ws, 34),
        "approver": cell(ws, 35),
        "envelope_p_min": number(cell(ws, 40, 2)),
        "envelope_p_max": number(cell(ws, 40, 3)),
        "envelope_t_min": number(cell(ws, 41, 2)),
        "envelope_t_max": number(cell(ws, 41, 3)),
        "envelope_dp_min": number(cell(ws, 42, 2)),
        "envelope_dp_max": number(cell(ws, 42, 3)),
        "envelope_gvf_min": number(cell(ws, 43, 2)),
        "envelope_gvf_max": number(cell(ws, 43, 3)),
        "envelope_wlr_min": number(cell(ws, 44, 2)),
        "envelope_wlr_max": number(cell(ws, 44, 3)),
    }
    source_row_count = 1
    source_sql = (
        "INSERT OR IGNORE INTO source_files "
        "(file_name,sha256,source_type,source_sheet,period_start,period_end,row_count) VALUES "
        f"({sql_value(file_name)},{sql_value(sha256)},{sql_value('calibracao_mpfm')},{sql_value('01_CAMPANHA')},"
        f"{sql_value(values['start_at'].strftime('%Y-%m-%d') if isinstance(values['start_at'], (datetime, date)) else None)},"
        f"{sql_value(values['end_at'].strftime('%Y-%m-%d') if isinstance(values['end_at'], (datetime, date)) else None)},"
        f"{source_row_count});"
    )
    statements.append(source_sql)
    source_ref = f"(SELECT id FROM source_files WHERE sha256={sql_value(sha256)} AND source_type='calibracao_mpfm')"

    columns = list(values.keys())
    statements.append(
        "INSERT OR REPLACE INTO calibration_campaigns "
        f"({','.join(columns)},source_file_id) VALUES "
        f"({','.join(sql_value(values[c]) for c in columns)},{source_ref});"
    )
    campaign_ref = f"(SELECT id FROM calibration_campaigns WHERE campaign_id={sql_value(campaign_id)})"
    return campaign_ref


def add_mpfm_rows(wb, campaign_ref: str, statements: list[str]):
    ws = wb["IN_01_MPFM"]
    for row in range(6, ws.max_row + 1):
        vals = [ws.cell(row=row, column=c).value for c in range(1, 18)]
        if not vals[1] or not vals[2]:
            continue
        columns = [
            "campaign_id", "condition", "timestamp", "use_flag", "duration_h", "quality",
            "pressure_barg", "temperature_c", "dp_kpa", "gvf_pct", "wlr_pct",
            "oil_uncorr_t", "gas_uncorr_t", "water_uncorr_t", "oil_corr_t", "gas_corr_t", "water_corr_t",
        ]
        row_values = [
            vals[1], vals[2], str(vals[3]).strip().upper() == "SIM", number(vals[4]), vals[5],
            number(vals[6]), number(vals[7]), number(vals[8]), number(vals[9]), number(vals[10]),
            number(vals[11]), number(vals[12]), number(vals[13]), number(vals[14]), number(vals[15]), number(vals[16]),
        ]
        statements.append(
            "INSERT OR REPLACE INTO calibration_mpfm_rows "
            f"(campaign_id,{','.join(columns[1:])}) VALUES "
            f"({campaign_ref},{sql_value(row_values[0])},"
            f"{','.join(sql_value(v) for v in row_values[1:])});"
        )


def add_separator_rows(wb, campaign_ref: str, statements: list[str]):
    ws = wb["IN_02_SEPARADOR"]
    for row in range(6, ws.max_row + 1):
        vals = [ws.cell(row=row, column=c).value for c in range(1, 17)]
        if not vals[1] or not vals[2]:
            continue
        columns = [
            "condition", "timestamp", "use_flag", "duration_h", "quality",
            "pressure_barg", "temperature_c", "oil_gv_line_m3", "oil_rho_coriolis_kgm3",
            "gas_mass_t", "water_mass_t", "gas_std_ksm3", "water_vol_m3", "source_ref",
        ]
        row_values = [
            vals[1], vals[2], str(vals[3]).strip().upper() == "SIM", number(vals[4]), vals[5],
            number(vals[6]), number(vals[7]), number(vals[8]), number(vals[9]),
            number(vals[10]), number(vals[11]), number(vals[12]), number(vals[13]), vals[14],
        ]
        statements.append(
            "INSERT OR REPLACE INTO calibration_separator_rows "
            f"(campaign_id,{','.join(columns)}) VALUES "
            f"({campaign_ref},{','.join(sql_value(v) for v in row_values)});"
        )


def add_lab_results(wb, campaign_ref: str, statements: list[str]):
    ws = wb["IN_03_LAB"]
    for row in range(6, ws.max_row + 1):
        vals = [ws.cell(row=row, column=c).value for c in range(1, 15)]
        if not isinstance(vals[3], (datetime, date)):
            continue
        columns = [
            "sample_id", "use_flag", "sampled_at", "sample_type", "bsw_pct",
            "rho_oil_std_kgm3", "rho_gas_std_kgsm3", "rho_water_std_kgm3", "fe", "rs",
            "method", "report_id", "status",
        ]
        row_values = [
            vals[1], str(vals[2]).strip().upper() == "SIM", vals[3], vals[4], number(vals[5]),
            number(vals[6]), number(vals[7]), number(vals[8]), number(vals[9]), number(vals[10]),
            vals[11], vals[12], vals[13],
        ]
        statements.append(
            "INSERT INTO calibration_lab_results "
            f"(campaign_id,{','.join(columns)}) VALUES "
            f"({campaign_ref},{','.join(sql_value(v) for v in row_values)});"
        )


def add_pvt_records(wb, campaign_ref: str, statements: list[str]):
    ws = wb["IN_04_PVT"]
    for row in range(6, ws.max_row + 1):
        vals = [ws.cell(row=row, column=c).value for c in range(1, 18)]
        if not vals[1]:
            continue
        columns = [
            "condition", "file", "sha256", "software", "version", "eos_model", "loaded_at", "approver",
            "input_oil_t", "input_gas_t", "input_water_t", "output_oil_t", "output_gas_t", "output_water_t",
            "pb_barg", "responded_at",
        ]
        row_values = [
            vals[1], vals[2], vals[3], vals[4], vals[5], vals[6], vals[7], vals[8],
            number(vals[9]), number(vals[10]), number(vals[11]), number(vals[12]), number(vals[13]), number(vals[14]),
            number(vals[15]), vals[16],
        ]
        statements.append(
            "INSERT OR REPLACE INTO calibration_pvt_records "
            f"(campaign_id,{','.join(columns)}) VALUES "
            f"({campaign_ref},{','.join(sql_value(v) for v in row_values)});"
        )


def add_k_applications(wb, campaign_ref: str, statements: list[str]):
    ws = wb["IN_05_K"]
    for row in range(6, ws.max_row + 1):
        vals = [ws.cell(row=row, column=c).value for c in range(1, 13)]
        if not vals[1]:
            continue
        columns = [
            "phase", "k_calculated", "k_approved", "k_applied", "applied_at",
            "responsible", "system", "config_version", "evidence_id", "status", "notes",
        ]
        row_values = [
            vals[1], number(vals[2]), number(vals[3]), number(vals[4]), vals[5],
            vals[6], vals[7], vals[8], vals[9], vals[10], vals[11],
        ]
        statements.append(
            "INSERT OR REPLACE INTO calibration_k_applications "
            f"(campaign_id,{','.join(columns)}) VALUES "
            f"({campaign_ref},{','.join(sql_value(v) for v in row_values)});"
        )


def add_uncertainty(wb, campaign_ref: str, statements: list[str]):
    ws = wb["IN_06_INCERTEZA"]
    for row in range(6, ws.max_row + 1):
        vals = [ws.cell(row=row, column=c).value for c in range(1, 11)]
        if not vals[1]:
            continue
        columns = [
            "condition", "u_mpfm_hc_pp", "u_mpfm_total_pp", "u_ref_hc_pp", "u_ref_total_pp",
            "k_mpfm", "k_ref", "source_version", "status",
        ]
        row_values = [
            vals[1], number(vals[2]), number(vals[3]), number(vals[4]), number(vals[5]),
            number(vals[6]), number(vals[7]), vals[8], vals[9],
        ]
        statements.append(
            "INSERT OR REPLACE INTO calibration_uncertainty "
            f"(campaign_id,{','.join(columns)}) VALUES "
            f"({campaign_ref},{','.join(sql_value(v) for v in row_values)});"
        )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--zip", type=Path, default=Path("Portal_MPFM_Riser_P4_v1.zip"))
    parser.add_argument("--output", type=Path, default=Path("drizzle/0003_calibration_riser_p4_campaign.sql"))
    args = parser.parse_args()
    if not args.zip.exists():
        raise SystemExit(f"Pacote não encontrado: {args.zip}")

    wb, file_name, sha256 = load_campaign_workbook(args.zip)
    statements = ["-- Gerado da campanha real de calibração MPFM (Riser P4); não editar manualmente."]
    campaign_ref = add_campaign(wb, file_name, sha256, statements)
    add_mpfm_rows(wb, campaign_ref, statements)
    add_separator_rows(wb, campaign_ref, statements)
    add_lab_results(wb, campaign_ref, statements)
    add_pvt_records(wb, campaign_ref, statements)
    add_k_applications(wb, campaign_ref, statements)
    add_uncertainty(wb, campaign_ref, statements)
    wb.close()

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text("\n".join(statements) + "\n", encoding="utf-8")
    print(f"{args.output}: {len(statements) - 1} comandos gerados de {file_name}")


if __name__ == "__main__":
    main()
