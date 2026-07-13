"""Profile the ANP Oil Linear export without copying raw rows into the portal.

The output contains provenance, aggregate quality checks, and auditable zero
classification rules. It intentionally excludes measurement values and serial
numbers so the generated artifact is safe to keep with the portal source.
"""

from __future__ import annotations

import argparse
import hashlib
import json
from collections import Counter, defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EXPECTED_SHEET = "Export"
REQUIRED_COLUMNS = (
    "Tag do Ponto Medição",
    "Início Período Medição",
    "Fim Período Medição",
    "Volume Bruto Corrigido (m3)",
    "Volume Bruto (m3)",
    "Volume Liquido (m3)",
)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def all_zero(values: list[Any]) -> bool:
    return all(is_number(value) and abs(float(value)) < 1e-12 for value in values)


def equal_numbers(left: Any, right: Any) -> bool:
    return is_number(left) and is_number(right) and abs(float(left) - float(right)) < 1e-9


def failure_flag(value: Any) -> bool:
    if value is True or value == 1:
        return True
    if isinstance(value, str):
        return value.strip().casefold() in {"1", "sim", "s", "true", "falha", "failure", "ativo"}
    return False


def iso(value: Any) -> str | None:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return None


def pct(count: int, total: int) -> float:
    return round((count / total * 100) if total else 0.0, 2)


def classify(row: dict[str, Any]) -> str:
    required_values = [row.get(column) for column in REQUIRED_COLUMNS]
    if any(value is None for value in required_values):
        return "missing_required"

    flags = [
        row.get("Estado BSW Falha"),
        row.get("Estado Inst Falha Temperatura"),
        row.get("Estado Instr Falha"),
        row.get("Estado Instrumento Falha"),
    ]
    if any(failure_flag(value) for value in flags):
        return "instrument_failure_flag"

    volumes = [
        row.get("Volume Bruto Corrigido (m3)"),
        row.get("Volume Bruto (m3)"),
        row.get("Volume Liquido (m3)"),
    ]
    volume_zero = all_zero(volumes)
    pt_zero = all_zero([row.get("Temperatura Fluido  (°C)"), row.get("Pressão Estática (kPa)")])
    totalizer_stable = equal_numbers(row.get("Vol Totalizado Inicio Prod"), row.get("Vol Totalizado Fim Prod"))

    if not volume_zero and pt_zero:
        return "flow_with_pt_zero"
    if volume_zero and totalizer_stable and pt_zero:
        return "zero_stable_totalizer_pt_zero"
    if volume_zero and totalizer_stable:
        return "zero_stable_totalizer_pt_present"
    if volume_zero:
        return "zero_other"
    return "normal_operating"


RULES = {
    "missing_required": {
        "label": "Campos obrigatórios ausentes",
        "status": "Ausente",
        "severity": "critical",
        "rule": "TAG, período e três volumes precisam estar preenchidos.",
    },
    "instrument_failure_flag": {
        "label": "Falha instrumental registrada",
        "status": "Em validação",
        "severity": "high",
        "rule": "Ao menos um estado de falha está explicitamente ativo na fonte.",
    },
    "flow_with_pt_zero": {
        "label": "Fluxo com P/T iguais a zero",
        "status": "Em validação",
        "severity": "high",
        "rule": "Há volume no período, mas temperatura e pressão são zero; investigar falha, substituição ou dado congelado.",
    },
    "zero_stable_totalizer_pt_zero": {
        "label": "Zero com totalizador estável e P/T zero",
        "status": "Não avaliável",
        "severity": "medium",
        "rule": "Candidato a sem fluxo, fora de operação ou ponto não selecionado; zero não é classificado como falha automaticamente.",
    },
    "zero_stable_totalizer_pt_present": {
        "label": "Zero com totalizador estável e P/T presentes",
        "status": "Em validação",
        "severity": "medium",
        "rule": "Candidato a sem fluxo ou ponto não selecionado; requer evidência operacional.",
    },
    "zero_other": {
        "label": "Zero sem causa determinada",
        "status": "Em validação",
        "severity": "high",
        "rule": "Volumes iguais a zero sem combinação conclusiva; revisar totalizadores, estados e alinhamento.",
    },
    "normal_operating": {
        "label": "Leitura com volume e P/T presentes",
        "status": "Derivado",
        "severity": "info",
        "rule": "Classificação estrutural; não equivale a conformidade metrológica.",
    },
}


def analyze(source_path: Path) -> dict[str, Any]:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    if EXPECTED_SHEET not in workbook.sheetnames:
        raise ValueError(f"Expected sheet {EXPECTED_SHEET!r}; found {workbook.sheetnames!r}")

    sheet = workbook[EXPECTED_SHEET]
    rows = sheet.iter_rows(values_only=True)
    headers = tuple(next(rows))
    missing_columns = sorted(set(REQUIRED_COLUMNS) - set(headers))
    if missing_columns:
        raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

    records: list[dict[str, Any]] = []
    for values in rows:
        record = dict(zip(headers, values))
        if not record.get("Tag do Ponto Medição"):
            continue
        records.append(record)

    total = len(records)
    key_counter: Counter[tuple[Any, Any, Any]] = Counter()
    tag_counter: Counter[str] = Counter()
    class_counter: Counter[str] = Counter()
    tag_classes: dict[str, Counter[str]] = defaultdict(Counter)
    starts: list[datetime] = []
    ends: list[datetime] = []

    for record in records:
        tag = str(record["Tag do Ponto Medição"]).strip()
        key = (tag, record.get("Início Período Medição"), record.get("Fim Período Medição"))
        key_counter[key] += 1
        tag_counter[tag] += 1
        category = classify(record)
        class_counter[category] += 1
        tag_classes[tag][category] += 1
        if isinstance(record.get("Início Período Medição"), datetime):
            starts.append(record["Início Período Medição"])
        if isinstance(record.get("Fim Período Medição"), datetime):
            ends.append(record["Fim Período Medição"])

    duplicate_rows = sum(count - 1 for count in key_counter.values() if count > 1)
    volume_zero = sum(
        1
        for record in records
        if all_zero(
            [
                record.get("Volume Bruto Corrigido (m3)"),
                record.get("Volume Bruto (m3)"),
                record.get("Volume Liquido (m3)"),
            ]
        )
    )
    pt_zero = sum(
        1
        for record in records
        if all_zero([record.get("Temperatura Fluido  (°C)"), record.get("Pressão Estática (kPa)")])
    )

    classifications = []
    for code, rule in RULES.items():
        count = class_counter[code]
        classifications.append({"code": code, **rule, "count": count, "rate": pct(count, total)})

    by_tag = []
    for tag in sorted(tag_counter):
        count = tag_counter[tag]
        zero_count = sum(value for code, value in tag_classes[tag].items() if code.startswith("zero_"))
        by_tag.append(
            {
                "tag": tag,
                "rows": count,
                "zeroRows": zero_count,
                "zeroRate": pct(zero_count, count),
                "flowWithPtZeroRows": tag_classes[tag]["flow_with_pt_zero"],
                "failureFlagRows": tag_classes[tag]["instrument_failure_flag"],
                "status": "Em validação" if zero_count or tag_classes[tag]["flow_with_pt_zero"] else "Derivado",
            }
        )

    return {
        "meta": {
            "dataset": "Óleo Linear",
            "sourceFile": source_path.name,
            "sourceSha256": sha256(source_path),
            "sheet": EXPECTED_SHEET,
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "status": "Em validação",
            "containsRawMeasurements": False,
        },
        "grain": "Uma linha por TAG e intervalo diário de medição",
        "profile": {
            "rows": total,
            "columns": len(headers),
            "tags": len(tag_counter),
            "measurementDateStart": iso(min(starts)) if starts else None,
            "measurementDateEnd": iso(max(starts)) if starts else None,
            "periodEndExclusive": iso(max(ends)) if ends else None,
        },
        "checks": {
            "compositeKey": "TAG + início do período + fim do período",
            "duplicateRows": duplicate_rows,
            "duplicateRate": pct(duplicate_rows, total),
            "volumeZeroRows": volume_zero,
            "volumeZeroRate": pct(volume_zero, total),
            "pressureTemperatureZeroRows": pt_zero,
            "pressureTemperatureZeroRate": pct(pt_zero, total),
        },
        "classifications": classifications,
        "byTag": by_tag,
        "limitations": [
            "As categorias de zero são hipóteses auditáveis, não causas operacionais confirmadas.",
            "A classificação final depende de alinhamento, estado operacional, seleção do ponto e evidência de falha/substituição.",
            "Nenhum indicador desta análise autoriza o uso do rótulo Conforme.",
        ],
    }


def render_markdown(result: dict[str, Any]) -> str:
    profile = result["profile"]
    checks = result["checks"]
    lines = [
        "# Qualidade de dados — Óleo Linear",
        "",
        f"Fonte: `{result['meta']['sourceFile']}` · aba `{result['meta']['sheet']}` · SHA-256 `{result['meta']['sourceSha256']}`",
        "",
        f"Grão: {result['grain']}.",
        "",
        "## Perfil",
        "",
        f"- {profile['rows']} linhas úteis e {profile['columns']} colunas;",
        f"- {profile['tags']} TAGs;",
        f"- datas de medição de {profile['measurementDateStart']} a {profile['measurementDateEnd']} (fim do último intervalo: {profile['periodEndExclusive']});",
        f"- {checks['duplicateRows']} duplicidades na chave composta ({checks['duplicateRate']}%).",
        "",
        "## Classificação inicial",
        "",
        "| Regra | Linhas | Taxa | Situação | Severidade |",
        "|---|---:|---:|---|---|",
    ]
    for item in result["classifications"]:
        lines.append(f"| {item['label']} | {item['count']} | {item['rate']}% | {item['status']} | {item['severity']} |")
    lines.extend(
        [
            "",
            "## Limitações",
            "",
            *[f"- {item}" for item in result["limitations"]],
            "",
        ]
    )
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("--json", type=Path)
    parser.add_argument("--report", type=Path)
    args = parser.parse_args()

    result = analyze(args.source.resolve())
    payload = json.dumps(result, ensure_ascii=False, indent=2)

    if args.json:
        args.json.parent.mkdir(parents=True, exist_ok=True)
        args.json.write_text(payload + "\n", encoding="utf-8")
    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(render_markdown(result), encoding="utf-8")
    if not args.json and not args.report:
        print(payload)


if __name__ == "__main__":
    main()
