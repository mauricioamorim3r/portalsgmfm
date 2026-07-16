"""Gera uma migração D1 somente a partir das fontes técnicas reais selecionadas."""

from __future__ import annotations

import argparse
import hashlib
import re
from collections import defaultdict
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook


def sql_value(value):
    if value is None or value == "":
        return "NULL"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, (int, float)):
        return str(value) if value == value else "NULL"
    return "'" + str(value).replace("'", "''") + "'"


def digest(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def iso_day(value) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = str(value or "").strip()
    return text[:10] if re.fullmatch(r"\d{4}-\d{2}-\d{2}.*", text) else ""


def number(value):
    if value in (None, ""):
        return None
    try:
        result = float(value)
        return result if result == result else None
    except (TypeError, ValueError):
        return None


def hour_number(value):
    if value in (None, "", "DAY"):
        return None
    if isinstance(value, time):
        return value.hour
    try:
        return int(float(value))
    except (TypeError, ValueError):
        match = re.search(r"\b(\d{1,2})\b", str(value))
        return int(match.group(1)) if match else None


def source_insert(path: Path, source_type: str, sheet: str, period_start: str, period_end: str, rows: int):
    sha = digest(path)
    sql = (
        "INSERT OR IGNORE INTO source_files "
        "(file_name,sha256,source_type,source_sheet,period_start,period_end,row_count) VALUES "
        f"({sql_value(path.name)},{sql_value(sha)},{sql_value(source_type)},{sql_value(sheet)},"
        f"{sql_value(period_start)},{sql_value(period_end)},{rows});"
    )
    ref = f"(SELECT id FROM source_files WHERE sha256={sql_value(sha)} AND source_type={sql_value(source_type)})"
    return sql, ref


def read_sheet(path: Path, sheet="Export"):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    result = []
    for row_number, row in enumerate(rows, 2):
        values = {header: row[index] if index < len(row) else None for index, header in enumerate(headers) if header}
        if any(value not in (None, "") for value in values.values()):
            result.append((row_number, values))
    wb.close()
    return result


def add_measurement_points(path: Path, statements: list[str]):
    rows = read_sheet(path)
    source_sql, source_ref = source_insert(path, "cadastro_pontos_medicao", "Export", "", "", len(rows))
    statements.append(source_sql)
    for _, row in rows:
        tag = str(row.get("TAG_PONTO_MEDICAO") or "").strip()
        if not tag:
            continue
        active = str(row.get("IND_ATIVO") or "").strip().lower() in {"1", "sim", "s", "true", "ativo"}
        values = [
            tag, row.get("COD_INSTALACAO"), row.get("FLUIDO"), row.get("TIPO_MEDICAO_PRINCIPAL"),
            row.get("TIPO_MEDICAO_SECUNDARIA"), row.get("TIPO_MEDIDOR"), row.get("NOM_LOCALIZACAO_PLANTA"),
            row.get("COMPUTADOR_VAZAO"), active,
        ]
        statements.append(
            "INSERT OR REPLACE INTO measurement_points "
            "(tag,installation_code,fluid,primary_measurement,secondary_measurement,meter_type,location,flow_computer,active,source_file_id) VALUES "
            f"({','.join(sql_value(v) for v in values)},{source_ref});"
        )


def add_wells(path: Path, statements: list[str]):
    rows = read_sheet(path)
    source_sql, source_ref = source_insert(path, "cadastro_pocos", "Export", "", "", len(rows))
    statements.append(source_sql)
    for _, row in rows:
        code = str(row.get("NÚMERO CADASTRO") or "").strip()
        if not code:
            continue
        values = [code, row.get("NOME POÇO ANP"), row.get("POÇO OPERADOR"), row.get("NOME CAMPO"), row.get("SITUAÇÃO ATUAL DO POÇO"), row.get("CATEGORIA")]
        statements.append(
            "INSERT OR REPLACE INTO wells (anp_code,anp_name,operator_name,field_name,status,category,source_file_id) VALUES "
            f"({','.join(sql_value(v) for v in values)},{source_ref});"
        )


def add_mpfm(path: Path, statements: list[str], groups: dict):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["BASE_UNICA_MES"]
    rows = ws.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    records = []
    days = []
    for row_number, row in enumerate(rows, 2):
        values = {header: row[index] if index < len(row) else None for index, header in enumerate(headers) if header}
        production_day = iso_day(values.get("ProductionDate"))
        tag = str(values.get("Tag") or "").strip()
        granularity = str(values.get("Granularity") or "").strip().title()
        if not production_day or not tag or granularity not in {"Hourly", "Daily", "Recon"}:
            continue
        days.append(production_day)
        record = (row_number, values, production_day, tag, granularity)
        records.append(record)
        key = (production_day, tag)
        if granularity == "Hourly":
            hour = hour_number(values.get("Hour"))
            if hour is not None:
                groups[key]["hours"].add(hour)
        elif granularity == "Daily":
            groups[key]["daily"] += 1
            groups[key]["daily_hc"] += number(values.get("MPFM corr HC (t)")) or 0
            groups[key]["daily_total"] += number(values.get("MPFM corr Total (t)")) or 0
    source_sql, source_ref = source_insert(path, "mpfm_base_unica", "BASE_UNICA_MES", min(days), max(days), len(records))
    statements.append(source_sql)
    for row_number, row, production_day, tag, granularity in records:
        values = [
            row_number, production_day, hour_number(row.get("Hour")), granularity, row.get("Origin"), row.get("Bank"),
            row.get("Loop"), row.get("Entity"), tag, row.get("Instrumento"), number(row.get("MPFM corr Gás (t)")),
            number(row.get("MPFM corr Óleo (t)")), number(row.get("MPFM corr HC (t)")),
            number(row.get("MPFM corr Água (t)")), number(row.get("MPFM corr Total (t)")),
            number(row.get("Pressão (barg)")), number(row.get("Temperatura (°C)")),
            str(row.get("IsOfficial") or "1").strip().lower() not in {"0", "false", "não", "nao"},
        ]
        statements.append(
            "INSERT OR IGNORE INTO mpfm_measurements "
            "(source_file_id,source_row,production_date,hour,granularity,origin,bank,loop,entity,tag,instrument,gas_t,oil_t,hc_t,water_t,total_t,pressure_barg,temperature_c,official) VALUES "
            f"({source_ref},{','.join(sql_value(v) for v in values)});"
        )
    wb.close()


def add_separator(path: Path, statements: list[str]):
    wb = load_workbook(path, read_only=True, data_only=True)
    definitions = {
        "separador oleo": {"phase": "oleo", "tag": "20FT0247", "pressure": 3, "temperature": 5, "volume": 10, "mass": 11, "flow": None},
        "separador gas": {"phase": "gas", "tag": "20FT0244", "pressure": 3, "temperature": 4, "volume": 8, "mass": 9, "flow": 12},
        "separador agua": {"phase": "agua", "tag": "20FT0251", "pressure": 3, "temperature": 4, "volume": 9, "mass": 10, "flow": None},
    }
    records = []
    days = []
    for sheet, definition in definitions.items():
        ws = wb[sheet]
        current_day = ""
        for row_number, row in enumerate(ws.iter_rows(values_only=True), 1):
            cells = list(row)
            marker = str(cells[2] or "") if len(cells) > 2 else ""
            match = re.search(r"Data:\s*(\d{4}-\d{2}-\d{2})", marker)
            if match:
                current_day = match.group(1)
                days.append(current_day)
                continue
            hour = hour_number(cells[2] if len(cells) > 2 else None)
            is_daily = str(cells[2] if len(cells) > 2 else "").strip().upper() == "DAY"
            if not current_day or (hour is None and not is_daily):
                continue
            records.append((sheet, row_number, current_day, None if is_daily else hour, definition, cells))
    source_sql, source_ref = source_insert(path, "separador_teste", "3 fases", min(days), max(days), len(records))
    statements.append(source_sql)
    for sheet, row_number, production_day, hour, definition, cells in records:
        def get(index, cells=cells):
            return number(cells[index]) if index is not None and index < len(cells) else None
        values = [sheet, row_number, production_day, hour, definition["phase"], definition["tag"], get(definition["pressure"]), get(definition["temperature"]), get(definition["volume"]), get(definition["mass"]), get(definition["flow"])]
        statements.append(
            "INSERT OR IGNORE INTO separator_measurements "
            "(source_file_id,source_sheet,source_row,production_date,hour,phase,tag,pressure,temperature_c,standard_volume,mass_t,flow_time_minutes) VALUES "
            f"({source_ref},{','.join(sql_value(v) for v in values)});"
        )
    wb.close()


def add_quality_issues(statements: list[str], groups: dict):
    for (production_day, tag), state in sorted(groups.items()):
        hours = state["hours"]
        if hours and len(hours) < 24:
            expected = set(range(1, 25)) if min(hours) >= 1 and max(hours) <= 24 else set(range(24))
            missing = ", ".join(f"{hour:02d}" for hour in sorted(expected - hours))
            statements.append(
                "INSERT INTO data_quality_issues (production_date,tag,issue_type,severity,details) VALUES "
                f"({sql_value(production_day)},{sql_value(tag)},'missing_hours','warn',{sql_value('Horas ausentes: ' + missing)});"
            )
        if state["daily"] and hours and state["daily_hc"] == 0 and state["daily_total"] == 0:
            statements.append(
                "INSERT INTO data_quality_issues (production_date,tag,issue_type,severity,details) VALUES "
                f"({sql_value(production_day)},{sql_value(tag)},'zero_daily_with_hourly','review','Daily zerado com dados horários presentes; causa não presumida.');"
            )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--downloads", type=Path, default=Path.home() / "Downloads")
    parser.add_argument("--cadastros", type=Path, default=Path.home() / "OneDrive/Documentos/Painel_Operador")
    parser.add_argument("--output", type=Path, default=Path("drizzle/0001_real_source_data.sql"))
    args = parser.parse_args()
    files = {
        "points": args.cadastros / "Pontos de Medição.xlsx",
        "wells": args.cadastros / "Poços.xlsx",
        "mpfm_jun": args.downloads / "MPFM_JUN_2026 (2).xlsx",
        "mpfm_jul": args.downloads / "MPFM_JUL_2026 (3).xlsx",
        "separator": args.downloads / "SEP_Dados_2026-06-01_a_2026-07-31 (1).xlsx",
    }
    missing = [str(path) for path in files.values() if not path.exists()]
    if missing:
        raise SystemExit("Fontes ausentes:\n" + "\n".join(missing))

    statements = ["-- Gerado de fontes reais; não editar manualmente.", "DELETE FROM data_quality_issues;"]
    groups = defaultdict(lambda: {"hours": set(), "daily": 0, "daily_hc": 0.0, "daily_total": 0.0})
    add_measurement_points(files["points"], statements)
    add_wells(files["wells"], statements)
    add_mpfm(files["mpfm_jun"], statements, groups)
    add_mpfm(files["mpfm_jul"], statements, groups)
    add_separator(files["separator"], statements)
    add_quality_issues(statements, groups)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text("\n".join(statements) + "\n", encoding="utf-8")
    print(f"{args.output}: {len(statements) - 2} comandos gerados de 5 fontes reais")


if __name__ == "__main__":
    main()
